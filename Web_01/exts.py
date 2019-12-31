import glob
import os
import zipfile
from shutil import copy, rmtree
import pandas as pd
from flask_sqlalchemy import SQLAlchemy
from openpyxl import load_workbook
from openpyxl.styles import Side, Border

db = SQLAlchemy()


# 创建目录
def creat_folder(folder_path):
    if not os.path.exists(folder_path):
        os.mkdir(folder_path)
        os.chmod(folder_path, os.O_RDWR)


# 删除文件
def delfile(path):
    #   read all the files under the folder
    fileNames = glob.glob(path + r'\*')
    for fileName in fileNames:
        try:
            #           delete file
            os.remove(fileName)
        except:
            rmtree(fileName)
            # try:
            #     #               delete empty folders
            #     os.rmdir(fileName)
            # except:
            #     #               Not empty, delete files under folders
            #     delfile(fileName)
            #     #               now, folders are empty, delete it
            #     os.rmdir(fileName)


def chengzhao_output_scores_tables(path):
    '''
    成招生成成绩单
    :param path: 模板所在目录
    :return:
    '''
    # todo  文件路径
    scores_path = f'{path}/scores.xlsx'
    templte_path = f'{path}/templte.xlsx'
    output_path = f'{path}/output/'
    creat_folder(output_path)

    # todo  提取数据
    df = pd.read_excel(io=scores_path, sheet_name=0, index_col='学号')
    # 数据结构 [['1601313030001', ['郭子铭', {'工程造价实训': '50', '数学': '30'}]], ['1601313030002', ...]]
    # 学号列表
    student_nums_ls = df.index.tolist()
    # 科目列表
    courses_ls = df.columns.tolist()[1:]
    # todo 获取学生数据
    data = []
    for i in student_nums_ls:
        ls = []
        student_name_ls = df.loc[i, ['姓名']].tolist()
        student_scores_ls = df.loc[i, '姓名':].tolist()[1:]
        scores_dict = dict(zip(courses_ls, student_scores_ls))
        student_name_ls.append(scores_dict)
        ls.append(str(i))
        ls.append(student_name_ls)
        data.append(ls)
    for i in range(len(data)):
        student_i = data[i]
        student_num = student_i[0]
        student_name = student_i[1][0]
        student_scores = student_i[1][1]
        path_output_ = '{}{}{}.xlsx'.format(output_path, student_num, student_name)
        copy(templte_path, path_output_)
        wb = load_workbook(path_output_)
        cols = wb['Sheet1'].max_column
        rows = wb['Sheet1'].max_row
        # 写入表格
        wb['Sheet1']['B6'] = student_num
        wb['Sheet1']['C6'] = '                      姓名：{}                        性别：男    '.format(student_name)
        # wb['Sheet1']['C{}'.format(rows)] = '                                            2019 年 3 月 1 日'
        for row in range(9, rows):
            key = wb['Sheet1'].cell(row=row, column=2).value
            if key in student_scores.keys():
                for col in range(5, cols + 1):
                    value = wb['Sheet1'].cell(row=row, column=col).value
                    if value == '*':
                        wb['Sheet1'].cell(row=row, column=col).value = student_scores[key]
                        # wb['Sheet1'].cell(row=row, column=col).border = border
        # 设置单元格样式
        thin = Side(border_style="thin", color='FF000000')
        border = Border(left=thin, right=thin, top=thin, bottom=thin)
        for row in range(7, rows + 1):
            for col in range(1, cols + 1):
                wb['Sheet1'].cell(row=row, column=col).border = border
        clean = Side(border_style=None)
        border_l = Border(left=clean, right=thin)
        border_r = Border(left=thin, right=clean)
        wb['Sheet1']['F{}'.format(rows - 1)].border = border_l
        wb['Sheet1']['J{}'.format(rows - 1)].border = border_l
        wb['Sheet1']['E{}'.format(rows - 1)].border = border_r
        wb['Sheet1']['I{}'.format(rows - 1)].border = border_r
        wb.save(path_output_)
        wb.close()


# 压缩文件
def to_zip(startdir):
    '''startdir:要压缩文件夹的绝对路径 '''
    file_news = startdir + '.zip'  # 压缩后文件夹的名字
    z = zipfile.ZipFile(file_news, 'w', zipfile.ZIP_DEFLATED)  # 参数一：文件夹名
    for dirpath, dirnames, filenames in os.walk(startdir):
        fpath = dirpath.replace(startdir, '')  # 这一句很重要，不replace的话，就从根目录开始复制
        fpath = fpath and fpath + os.sep or ''  # 这句话理解我也点郁闷，实现当前文件夹以及包含的所有文件的压缩
        for filename in filenames:
            z.write(os.path.join(dirpath, filename), fpath + filename)
    print('压缩成功')
    z.close()
    return file_news


def part_print_do(path='考场情况表.xls', my_file='data.txt'):
    '''
    打印门贴签到表
    :param path:考场情况表路径
    :param my_file:生成文件路径
    :return:
    '''
    df = pd.read_excel(io=path, sheet_name='排考', dtype=str)
    exam_room_nums = df['考场号'].values.tolist()

    ls = []
    for exam_room_num in exam_room_nums:
        # first_num = str(exam_room_num)[0]
        print_num = int(str(exam_room_num)[1:])
        data = 'Set sh = Worksheets("sheet1"): sh.PrintOut {}, {}, , False'.format(print_num, print_num)
        ls.append(data)

    # todo 写入文件
    # 删除已有文件
    if os.path.exists(my_file):  # 如果文件存在
        # 删除文件，可使用以下两种方法。
        os.remove(my_file)  # 则删除
        # os.unlink(my_file)

    # 写入路径并输出
    def output_data(ls):
        with open(my_file, 'a+', encoding="utf-8") as f:
            # f.write('# -*- coding: utf-8 -*-\n')
            # f.write('# coding=gbk\n\n')
            f.write('Sub print1() \'定义打印过程' + '\n')
            f.write('Dim sh As Worksheet \'声明打印变量' + '\n')
            for i in ls:
                f.write(i + '\n')
            f.write('End Sub')

    output_data(ls)


'''
开放教育排课-----------------------------------------------------------------------------------------------
'''


def get_dict_class_level(path='class_info.xlsx'):
    '''
    获取优先级字典
    :param path:
    :return:
    '''
    df = pd.read_excel(io=path, sheet_name='优先级')
    ls_class_name = df['课程名称'].tolist()
    ls_class_level = df['优先级'].tolist()
    dict_class_level = dict(zip(ls_class_name, ls_class_level))
    return dict_class_level


def get_dict_classroom(path='class_info.xlsx', name='开放教室'):
    '''
    获取教室字典
    :param path:
    :param name:
    :return:
    '''
    df = pd.read_excel(io=path, sheet_name=name)
    ls_class_num = df['教室'].tolist()
    ls_class_vol = df['容量'].tolist()
    dict_class_room = dict(zip(ls_class_num, ls_class_vol))
    return dict_class_room


def get_dict_date(path='class_info.xlsx', name='开放上课时间'):
    '''
    获取上课日期
    :param path:
    :param name:
    :return:
    '''
    df = pd.read_excel(io=path, sheet_name=name)
    ls_time_part = df['时间段'].tolist()
    ls_date = df['日期'].tolist()
    dict_date = dict(zip(ls_time_part, ls_date))
    return dict_date


def get_dict_time(path='class_info.xlsx', name='开放上课时间'):
    '''
    获取上课时间
    :param path:
    :param name:
    :return:
    '''
    df = pd.read_excel(io=path, sheet_name=name)
    ls_time_part = df['时间段'].tolist()
    ls_time = df['时间'].tolist()
    dict_time = dict(zip(ls_time_part, ls_time))
    return dict_time


# dict_class_level = get_dict_class_level()
# dict_classroom = get_dict_classroom()
# dict_date = get_dict_date()
# dict_time = get_dict_time()


def get_1_time(courses_dict, set_classes, set_teachers, CLASSROOM_NUM, dict_class_level):
    '''
    # 获取一个时间段数据
    :param courses_dict: 课程所对应的班级字典列表
    :param set_classes: 班级集合（总表）
    :param ls_courses: 课程列表（每一次排课需要传入刨除已成功的课程）
    :return:data_ls(一次排课数据), have_courses_set(已被安排的课程)
    '''
    # 结果数据列表
    data_ls = []
    # 已选课程集合
    have_courses_set = set()
    # 用来记录课程名称
    classes_ls_ = []
    # 初始化计数器
    COUNTER = 1
    # 从优先级课程字典取出按顺序取出课程
    for course, _ in dict_class_level.items():
        # print(course)
        if not COUNTER == 0:
            classes_ls = courses_dict[course]  # 课程所对应的班级字典列表
            # 获取班级名称集合
            classname_set = set()
            # 获取任课教师集合
            tescher_set = set()
            for class_dict in classes_ls:
                classname = list(class_dict.keys())[0]
                classname_set.add(classname)
                for k, v in class_dict.items():
                    tescher_set.add(v[0])

            # 用集合判断班级名称是否在班级列表（总表）中
            if classname_set < set_classes and tescher_set < set_teachers:
                have_courses_set.add(course)
                for class_dict in classes_ls:
                    for k, v in class_dict.items():
                        ls = []
                        ls.append(course)  # 添加课程名称
                        ls.append(k)  # 添加班级名称
                        for i in v:
                            ls.append(i)  # 添加其他信息
                        data_ls.append(ls)
                        # 记录已编排课程名称
                        classes_ls_.append(course)
                        # 列表去重
                        classes_ls_ = list(set(classes_ls_))
                        # print(classes_ls_)
                        # print(len(data_all_ls))
                set_classes = set_classes - classname_set
                set_teachers = set_teachers - tescher_set
                # dict_class_level.pop(course)  # 从优先级课程列表弹出课程
        else:
            print('教室满了')
            break
        COUNTER = len(classes_ls_) % CLASSROOM_NUM
        # print(classes_ls_)
        # print(COUNTER)
    return data_ls, have_courses_set


def class_arrange_do(path, CLASSROOM_NUM, class_info_path, output_path):
    # df = pd.read_excel(io=path, sheet_name='自开开课一览表')
    df = pd.read_excel(io=path, sheet_name='网授开课一览表')
    col_names_ls = df.columns.tolist()  # 获取标题行列表
    col_names_ls.insert(0, '时间段')
    # col_names_ls.insert(1, '教室')

    # todo 提取数据结构

    # todo 唯一值列表
    ls_teachers = df['任课教师'].unique().tolist()
    ls_courses = df['课程名称'].unique().tolist()
    ls_classes = df['专业班级'].unique().tolist()

    # todo 建立字典

    def get_courses_dict():
        '''
        以课程为键，建立"课程:班级"字典
        :return: 返回"课程:班级"字典
        '''
        courses_dict = dict()  # 创建课程空字典
        for course in ls_courses:  # 从课程列表中取课程
            ls = []  # 创建一个空列表，用来存放班级
            for i in range(df.shape[0]):  # 从表中提取数据
                if df.iloc[i, 0] == course:  # 如果找到相同的课程名称
                    d = dict()  # 创建一个空字典用来存放"班级:其它"信息字典

                    # 把其它信息作为一个列表
                    data_ls = []
                    for j in range(2, df.shape[1]):
                        data_ls.append(df.iloc[i, j])

                    d.update({df.iloc[i, 1]: data_ls})  # 加入字典
                    # 把班级----其他信息字典加作为元素入列表
                    ls.append(d)  # 把对应的专业班级添加进列表

            courses_dict.update({course: ls})  # 得到"课程:班级"字典
        return courses_dict

    # todo 思路：通过课程，找到管理班（包含任课教师等所有信息）;
    # todo 从课程列表取出一门课;
    # todo 该课程对应的值也就是班级列表中的每一个班在班级列表（总表）中进行查询;
    # todo 如果每个值都找到了，把数据以列表的形式写入结果列表;
    # todo 同时从课程总列表中弹出这个课程，从班级列表（总表）中弹出找到的班级;

    # 获取班级、课程名称、任课教师集合
    set_classes = set(ls_classes)
    set_courses = set(ls_courses)
    set_teachers = set(ls_teachers)
    courses_dict = get_courses_dict()  # 调用函数创建"课程:班级"字典

    # todo 程序执行部分
    # 数据初始化
    i = 1
    data_all_ls = []
    dict_class_level = get_dict_class_level(class_info_path)  # 获取优先级字典
    dict_classroom = list(get_dict_classroom(class_info_path).keys())  # 获取教室字典
    dict_date = get_dict_date(class_info_path)  # 获取上课日期
    dict_time = get_dict_time(class_info_path)  # 获取上课时间
    while len(set_courses) > 0:
        # print(len(set_courses))  # 如果课程长度大于零，说明课程没有被排完
        # todo 设置教室数量
        # CLASSROOM_NUM = 14

        ls_courses = list(set_courses)
        # 获取排课列表
        data_ls, have_courses_set = get_1_time(courses_dict, set_classes, set_teachers, CLASSROOM_NUM, dict_class_level)
        # 添加时间段序号,及教室号
        class_room_num = 0
        for data in data_ls:
            data.insert(0, str(i))
            # try:
            #     data.insert(1, list(dict_kf_classroom.keys())[class_room_num])
            #     # print(data)
            # except:
            #     break
            data_all_ls.append(data)
            class_room_num += 1
        # 从课程中减去已排课程
        set_courses = set_courses - have_courses_set
        # 从优先级课程字典中弹出已排课程
        for have_course in list(have_courses_set):
            dict_class_level.pop(have_course)
        i += 1

    # 插入教室
    j = 0
    data_all_ls[0].insert(1, dict_classroom[0])
    for i in range(1, len(data_all_ls)):
        if data_all_ls[i][1] == data_all_ls[i - 1][2]:
            data_all_ls[i].insert(1, dict_classroom[j % len(dict_classroom)])
        else:
            j += 1
            data_all_ls[i].insert(1, dict_classroom[j % len(dict_classroom)])
    col_names_ls.insert(1, '教室')

    # 插入日期、时间
    for i in range(len(data_all_ls)):
        t_key = data_all_ls[i][0]  # 总表时间段作为键值
        data_all_ls[i].insert(2, dict_date[int(t_key)])
        data_all_ls[i].insert(3, dict_time[int(t_key)])
    col_names_ls.insert(2, '日期')
    col_names_ls.insert(3, '时间')
    # todo 写入Excel
    data_all = pd.DataFrame(data_all_ls, columns=col_names_ls)
    writer = pd.ExcelWriter(output_path + '排课结果.xlsx')
    data_all.to_excel(writer, '排课结果', index=False)
    writer.save()
